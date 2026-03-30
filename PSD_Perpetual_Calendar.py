"""
PSD Perpetual School Calendar Builder
Generates PSD_Perpetual_Calendar.xlsx using openpyxl.
Usage: python PSD_Perpetual_Calendar.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Color
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = r"C:\Users\johnw\Documents\GitHub\AI\PSD_Perpetual_Calendar.xlsx"

# ---------------------------------------------------------------------------
# Helper: thin border side
# ---------------------------------------------------------------------------
thin = Side(style="thin")
medium = Side(style="medium")

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def medium_border():
    return Border(left=medium, right=medium, top=medium, bottom=medium)


# ---------------------------------------------------------------------------
# Color constants
# ---------------------------------------------------------------------------
DARK_BLUE    = "1F4E79"
LIGHT_BLUE   = "D6E4F0"
WHITE        = "FFFFFF"
YELLOW       = "FFFF00"
BLUE_INPUT   = "0070C0"
RED          = "C00000"
GREEN        = "375623"
DARK_ORANGE  = "C55A11"
PURPLE       = "7030A0"
GRAY         = "595959"

# ---------------------------------------------------------------------------
# Holiday data: (formula, event, notes, type, description)
# ---------------------------------------------------------------------------
HOLIDAYS = [
    # Row 2
    (
        "=DATE('School Year'!$A$1,9,1)+MOD(3-WEEKDAY(DATE('School Year'!$A$1,9,1)),7)",
        "First Day of School", "TK–12", "First Day",
        "First Tuesday of September (fall year)"
    ),
    # Row 3
    (
        "=DATE('School Year'!$A$1,9,1)+MOD(2-WEEKDAY(DATE('School Year'!$A$1,9,1)),7)",
        "No School – Labor Day", "TK–12", "Federal Holiday",
        "1st Monday of September (fall year)"
    ),
    # Row 4
    (
        "=IF(WEEKDAY(DATE('School Year'!$A$1,11,11))=7,"
        "DATE('School Year'!$A$1,11,10),"
        "IF(WEEKDAY(DATE('School Year'!$A$1,11,11))=1,"
        "DATE('School Year'!$A$1,11,12),"
        "DATE('School Year'!$A$1,11,11)))",
        "No School – Veterans Day", "TK–12", "Federal Holiday",
        "November 11 (fixed); if Saturday → observed Friday Nov 10; if Sunday → observed Monday Nov 12"
    ),
    # Row 5
    (
        "=DATE('School Year'!$A$1,11,1)+MOD(6-WEEKDAY(DATE('School Year'!$A$1,11,1)),7)+14",
        "No School – Conferences", "Grades 6–8", "Conferences",
        "3rd Friday of November (fall year)"
    ),
    # Row 6
    (
        "=DATE('School Year'!$A$1,11,1)+MOD(5-WEEKDAY(DATE('School Year'!$A$1,11,1)),7)+20",
        "Early Release", "TK–12", "Early Release",
        "Wednesday before Thanksgiving: 4th Thursday of November minus 1 day"
    ),
    # Row 7
    (
        "=DATE('School Year'!$A$1,11,1)+MOD(5-WEEKDAY(DATE('School Year'!$A$1,11,1)),7)+21",
        "No School – Thanksgiving", "TK–12", "Federal Holiday",
        "4th Thursday of November (Thanksgiving Day, fall year)"
    ),
    # Row 8
    (
        "=DATE('School Year'!$A$1,11,1)+MOD(5-WEEKDAY(DATE('School Year'!$A$1,11,1)),7)+22",
        "No School – Thanksgiving Break", "TK–12", "Break",
        "Day after Thanksgiving (4th Thursday of November plus 1 day)"
    ),
    # Row 9
    (
        "=DATE('School Year'!$A$1,11,1)+MOD(5-WEEKDAY(DATE('School Year'!$A$1,11,1)),7)+25",
        "No School – Trimester Break", "TK–12", "Break",
        "Monday after Thanksgiving: 4th Thursday of November plus 4 days"
    ),
    # Row 10
    (
        "=DATE('School Year'!$A$1,12,1)+MOD(6-WEEKDAY(DATE('School Year'!$A$1,12,1)),7)+7",
        "No School – Conferences", "TK–5", "Conferences",
        "2nd Friday of December (fall year)"
    ),
    # Row 11
    (
        "=DATE('School Year'!$A$1,12,1)+MOD(5-WEEKDAY(DATE('School Year'!$A$1,12,1)),7)+13",
        "Early Release", "TK–12", "Early Release",
        "Day before Winter Break: 3rd Thursday of December minus 1 day (Wednesday)"
    ),
    # Row 12
    (
        "=DATE('School Year'!$A$1,12,1)+MOD(5-WEEKDAY(DATE('School Year'!$A$1,12,1)),7)+14",
        "No School – Winter Break Begins", "TK–12", "Break",
        "3rd Thursday of December (fall year); first day of Winter Break"
    ),
    # Row 13
    (
        "=DATE('School Year'!$A$1,12,31)",
        "No School – Winter Break Ends", "TK–12", "Break",
        "December 31 (fixed); last day of calendar year / Winter Break"
    ),
    # Row 14
    (
        "=DATE('School Year'!$A$1+1,1,1)",
        "No School – New Year's Day", "TK–12", "Federal Holiday",
        "January 1 (fixed, spring year)"
    ),
    # Row 15
    (
        "=DATE('School Year'!$A$1+1,1,1)+MOD(2-WEEKDAY(DATE('School Year'!$A$1+1,1,1)),7)+14",
        "No School – Martin Luther King Jr. Day", "TK–12", "Federal Holiday",
        "3rd Monday of January (spring year)"
    ),
    # Row 16
    (
        "=DATE('School Year'!$A$1+1,2,0)-MOD(WEEKDAY(DATE('School Year'!$A$1+1,2,0))-6+7,7)",
        "No School – Semester Break", "TK–12", "Break",
        "Last Friday of January (spring year): last day of January minus days-since-Friday"
    ),
    # Row 17
    (
        "=DATE('School Year'!$A$1+1,2,1)+MOD(2-WEEKDAY(DATE('School Year'!$A$1+1,2,1)),7)+14",
        "No School – Presidents' Day", "TK–12", "Federal Holiday",
        "3rd Monday of February (spring year)"
    ),
    # Row 18
    (
        "=DATE('School Year'!$A$1+1,2,1)+MOD(2-WEEKDAY(DATE('School Year'!$A$1+1,2,1)),7)+21",
        "Snow Make-Up Day (if needed)", "TK–12", "Make-Up Day",
        "Monday after Presidents' Day: Presidents' Day (3rd Monday of Feb) plus 7 days"
    ),
    # Row 19
    (
        "=DATE('School Year'!$A$1+1,3,1)+MOD(4-WEEKDAY(DATE('School Year'!$A$1+1,3,1)),7)+14",
        "No School – Trimester Break", "TK–12", "Break",
        "3rd Wednesday of March (spring year)"
    ),
    # Row 20
    (
        "=DATE('School Year'!$A$1+1,3,1)+MOD(6-WEEKDAY(DATE('School Year'!$A$1+1,3,1)),7)+21",
        "No School – Conferences", "TK–5", "Conferences",
        "4th Friday of March (spring year)"
    ),
    # Row 21
    (
        "=DATE('School Year'!$A$1+1,4,1)+MOD(2-WEEKDAY(DATE('School Year'!$A$1+1,4,1)),7)-3",
        "Early Release", "TK–12", "Early Release",
        "Friday before Spring Break: First Monday of April minus 3 days"
    ),
    # Row 22
    (
        "=DATE('School Year'!$A$1+1,4,1)+MOD(2-WEEKDAY(DATE('School Year'!$A$1+1,4,1)),7)",
        "No School – Spring Break Begins", "TK–12", "Break",
        "First Monday of April (spring year); first day of Spring Break"
    ),
    # Row 23
    (
        "=DATE('School Year'!$A$1+1,4,1)+MOD(2-WEEKDAY(DATE('School Year'!$A$1+1,4,1)),7)+4",
        "No School – Spring Break Ends", "TK–12", "Break",
        "First Friday of April: Spring Break start (First Monday of April) plus 4 days"
    ),
    # Row 24
    (
        "=DATE('School Year'!$A$1+1,5,0)-MOD(WEEKDAY(DATE('School Year'!$A$1+1,5,0))-6+7,7)",
        "No School – Conferences", "Grades 6–8", "Conferences",
        "Last Friday of April (spring year)"
    ),
    # Row 25
    (
        "=DATE('School Year'!$A$1+1,6,0)-MOD(WEEKDAY(DATE('School Year'!$A$1+1,6,0))-2+7,7)",
        "No School – Memorial Day", "TK–12", "Federal Holiday",
        "Last Monday of May (spring year): last day of May minus days-since-Monday"
    ),
    # Row 26
    (
        "=DATE('School Year'!$A$1+1,6,1)+MOD(5-WEEKDAY(DATE('School Year'!$A$1+1,6,1)),7)+14",
        "Last Day of School (Early Release)", "TK–12", "Last Day",
        "3rd Thursday of June (spring year)"
    ),
    # Row 27
    (
        "=IF(WEEKDAY(DATE('School Year'!$A$1+1,6,19))=7,"
        "DATE('School Year'!$A$1+1,6,18),"
        "IF(WEEKDAY(DATE('School Year'!$A$1+1,6,19))=1,"
        "DATE('School Year'!$A$1+1,6,20),"
        "DATE('School Year'!$A$1+1,6,19)))",
        "No School – Juneteenth Observance", "TK–12", "Federal Holiday",
        "June 19 (Juneteenth, fixed); if Saturday → observed Friday June 18; if Sunday → observed Monday June 20"
    ),
    # Row 28
    (
        "=A26-WEEKDAY(A26)+9",
        "Snow Make-Up Day (if needed)", "TK–12", "Make-Up Day",
        "First Monday after Last Day of School: Last Day minus WEEKDAY(Last Day) plus 9"
    ),
]


def type_font_color(type_str):
    """Return (hex_color, bold, italic) for the Type column based on event type."""
    mapping = {
        "Federal Holiday": (RED,        False, False),
        "First Day":       (GREEN,      True,  False),
        "Last Day":        (GREEN,      True,  False),
        "Break":           (DARK_ORANGE,False, False),
        "Conferences":     (PURPLE,     False, False),
        "Early Release":   (DARK_BLUE,  False, False),
        "Make-Up Day":     (GRAY,       False, True),
    }
    return mapping.get(type_str, ("000000", False, False))


def build_workbook():
    wb = Workbook()

    # -----------------------------------------------------------------------
    # Tab 1: School Year
    # -----------------------------------------------------------------------
    ws_sy = wb.active
    ws_sy.title = "School Year"

    # A1 — fall year input
    ws_sy["A1"] = 2026
    ws_sy["A1"].font = Font(name="Arial", size=20, bold=True, color=BLUE_INPUT)
    ws_sy["A1"].fill = PatternFill("solid", fgColor=YELLOW)
    ws_sy["A1"].border = thin_border()

    ws_sy["B1"] = "School Year (Fall Year — enter 4-digit start year)"
    ws_sy["B1"].font = Font(name="Arial", size=11)

    # A2 — spring year formula
    ws_sy["A2"] = "=A1+1"
    ws_sy["A2"].font = Font(name="Arial", size=14, color="000000")
    ws_sy["A2"].border = thin_border()

    ws_sy["B2"] = "Spring Year (auto-calculated)"
    ws_sy["B2"].font = Font(name="Arial", size=11)

    # A3–A5 instructions
    instructions = [
        "Instructions: Change the value in A1 to the desired fall year (e.g., 2027 for the 2027-28 school year).",
        "All holiday dates in the 'Holidays' tab will automatically recalculate based on that year.",
        "The yellow cell (A1) is the only cell you need to edit.",
    ]
    for i, text in enumerate(instructions, start=3):
        ws_sy.cell(row=i, column=1, value=text)
        ws_sy.cell(row=i, column=1).font = Font(name="Arial", size=10)

    ws_sy.column_dimensions["A"].width = 10
    ws_sy.column_dimensions["B"].width = 58

    # -----------------------------------------------------------------------
    # Tab 2: Holidays
    # -----------------------------------------------------------------------
    ws_h = wb.create_sheet("Holidays")

    headers = ["Date", "Event Name", "Grade Level / Notes", "Type",
               "Perpetual Calendar Formula"]
    header_fill   = PatternFill("solid", fgColor=DARK_BLUE)
    header_font   = Font(name="Arial", size=11, bold=True, color=WHITE)
    header_align  = Alignment(horizontal="center", vertical="center")

    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws_h.cell(row=1, column=col_idx, value=hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    ws_h.row_dimensions[1].height = 20

    # Column widths
    col_widths = [14, 36, 16, 18, 65]
    for col_idx, width in enumerate(col_widths, start=1):
        ws_h.column_dimensions[get_column_letter(col_idx)].width = width

    date_num_fmt = "MM/DD/YYYY"

    white_fill     = PatternFill("solid", fgColor=WHITE)
    lightblue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    for row_offset, (formula, event, notes, htype, desc) in enumerate(HOLIDAYS):
        excel_row = row_offset + 2          # data starts at row 2
        is_even   = (excel_row % 2 == 0)   # even excel rows get light-blue bg
        bg_fill   = lightblue_fill if is_even else white_fill

        hex_color, bold, italic = type_font_color(htype)

        # Col A — date formula
        cell_a = ws_h.cell(row=excel_row, column=1, value=formula)
        cell_a.number_format = date_num_fmt
        cell_a.font          = Font(name="Arial", size=10)
        cell_a.fill          = bg_fill
        cell_a.alignment     = Alignment(horizontal="center", vertical="center")

        # Col B — event name
        cell_b = ws_h.cell(row=excel_row, column=2, value=event)
        cell_b.font      = Font(name="Arial", size=10)
        cell_b.fill      = bg_fill
        cell_b.alignment = Alignment(vertical="center")

        # Col C — notes
        cell_c = ws_h.cell(row=excel_row, column=3, value=notes)
        cell_c.font      = Font(name="Arial", size=10)
        cell_c.fill      = bg_fill
        cell_c.alignment = Alignment(horizontal="center", vertical="center")

        # Col D — type (color-coded)
        cell_d = ws_h.cell(row=excel_row, column=4, value=htype)
        cell_d.font      = Font(name="Arial", size=10, bold=bold, italic=italic,
                                color=hex_color)
        cell_d.fill      = bg_fill
        cell_d.alignment = Alignment(horizontal="center", vertical="center")

        # Col E — description (wrap text)
        cell_e = ws_h.cell(row=excel_row, column=5, value=desc)
        cell_e.font      = Font(name="Arial", size=10)
        cell_e.fill      = bg_fill
        cell_e.alignment = Alignment(wrap_text=True, vertical="center")

        ws_h.row_dimensions[excel_row].height = 30

    # Borders: bold outer, thin inner
    last_data_row = len(HOLIDAYS) + 1  # row 28
    last_col      = 5

    for r in range(1, last_data_row + 1):
        for c in range(1, last_col + 1):
            cell = ws_h.cell(row=r, column=c)
            top_side    = medium if r == 1            else thin
            bottom_side = medium if r == last_data_row else thin
            left_side   = medium if c == 1            else thin
            right_side  = medium if c == last_col     else thin
            cell.border = Border(
                top=top_side, bottom=bottom_side,
                left=left_side, right=right_side
            )

    # Freeze top row
    ws_h.freeze_panes = "A2"

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    build_workbook()
